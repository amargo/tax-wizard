import os
import sys
import argparse
import pandas as pd
import xml.etree.ElementTree as ET

from datetime import datetime, timedelta
from io import StringIO
from openpyxl.utils import get_column_letter
from zeep import Client, Settings
from excel_config import SHEET_FORMAT_CONFIGS, HUF_FORMAT, NUMBER_FORMAT

STATEMENT_PREFIX = "[statement][transactions]"
exchange_rate_cache = { }


class MNBExchangeService:
    """Az MNB árfolyam szolgáltatásával kapcsolatos műveletek, cache-eléssel."""
    def __init__(self, wsdl_url="http://www.mnb.hu/arfolyamok.asmx?wsdl", strict=False):
        settings = Settings(strict=strict, xml_huge_tree=True)
        self.client = Client(wsdl_url, settings=settings)

    def get_exchange_rate(self, date: datetime, currency: str) -> float:
        """Lekéri az adott dátum (fallback esetén az előző nap) alapján a deviza árfolyamát.
        Ha a deviza HUF, visszatér 1.0-vel. A lekérdezést cache-eli.
        """
        key = (date.strftime("%Y-%m-%d"), currency.upper())
        if key in exchange_rate_cache:
            return exchange_rate_cache[key]
        if currency.upper() == "HUF":
            exchange_rate_cache[key] = 1.0
            return 1.0

        start_date_str = (date - timedelta(days=5)).strftime("%Y-%m-%d")
        end_date_str = date.strftime("%Y-%m-%d")
        request_data = {
            'startDate': start_date_str,
            'endDate': end_date_str,
            'currencyNames': currency.upper()
        }
        try:
            response_xml = self.client.service.GetExchangeRates(**request_data)
        except Exception as e:
            print("Hiba a GetExchangeRates metódus hívásakor:", e)
            return None

        try:
            root = ET.fromstring(response_xml)
        except Exception as e:
            print(f"XML feldolgozási hiba: {e}")
            return None

        days = []
        for day in root.findall("Day"):
            day_date = day.attrib.get("date")
            try:
                parsed_date = datetime.strptime(day_date, "%Y-%m-%d")
                days.append((parsed_date, day))
            except Exception as ex:
                print(f"Hiba a dátum értelmezésekor: {day_date} -> {ex}")

        if not days:
            print("Nem található 'Day' elem a válaszban.")
            return None

        requested_date_str = date.strftime("%Y-%m-%d")
        # Megpróbáljuk megtalálni a kért dátumhoz tartozó napot,
        # ha nem, akkor a legutolsó elérhető napot választjuk.
        chosen_day = None
        for d, day in days:
            if d.strftime("%Y-%m-%d") == requested_date_str:
                chosen_day = day
                break
        if not chosen_day:
            chosen_day = max(days, key=lambda x: x[0])[1]

        for rate in chosen_day.findall("Rate"):
            if rate.attrib.get("curr", "").upper() == currency.upper():
                try:
                    value = float(rate.text.replace(",", "."))
                    exchange_rate_cache[key] = value
                    return value
                except Exception as e:
                    print(f"Az árfolyam érték konvertálása sikertelen ({rate.text}): {e}")
                    return None

        print(f"A {currency} árfolyam nem található a válaszban.")
        return None

    def convert_to_huf(self, amount: float, date: datetime, currency: str) -> float:
        """Átváltja az adott összeget forintra a lekérdezett árfolyammal."""
        rate = self.get_exchange_rate(date, currency)
        if rate is None:
            print(f"Nincs árfolyam adat {currency} esetén {date.strftime('%Y-%m-%d')}.")
            return None
        return amount * rate


class LightyearProcessor:
    """Lightyear CSV tranzakciós adatok feldolgozása."""
    def __init__(self, csv_file: str):
        self.csv_file = csv_file
        self.df = pd.read_csv(csv_file, parse_dates=["Date"], dayfirst=True)
        self.df["Type"] = self.df["Type"].astype(str).str.strip()
        self.df["CCY"] = self.df["CCY"].astype(str).str.strip()
        self.exchange_service = MNBExchangeService()

    def process(self):
        """Feldolgozza a CSV-t és DataFrame-eket készít:
           - realized_df: Realizált (eladott) pozíciók
           - open_df: Nyitott pozíciók
           - interest_df: Kamatjövedelem
           - dividend_df: Osztalékjövedelem
        """
        trades = self.df[
            self.df["Type"].isin(["Buy", "Sell", "Distribution"]) &
            (self.df["Ticker"].notnull()) &
            (self.df["Ticker"] != "")
        ]
        interests = self.df[self.df["Type"] == "Interest"]
        dividends = self.df[self.df["Type"] == "Dividend"]

        realized_df, open_df = self._process_trades(trades)
        interest_df = self._process_income(interests)
        dividend_df = self._process_income(dividends)
        return realized_df, open_df, interest_df, dividend_df

    def _process_trades(self, trades: pd.DataFrame):
        realized_list = []
        open_positions_list = []

        for (ticker, ccy), group in trades.groupby(["Ticker", "CCY"]):
            # Szűrés: vásárlások (Buy, Distribution) és eladások (Sell)
            buy_rows = group[group["Type"].isin(["Buy", "Distribution"])]
            sell_rows = group[group["Type"] == "Sell"]

            # Összegzés idegen pénznemben
            total_buy_fc = buy_rows["Net Amt."].sum()
            total_sell_fc = sell_rows["Net Amt."].sum() if not sell_rows.empty else 0

            # Vásárlások HUF-ban: minden sorra külön átváltás (vásárlás napján érvényes árfolyam)
            total_buy_huf = self._sum_in_huf(buy_rows, ccy)

            # Eladások HUF-ban: minden eladási sorra külön átváltás (eladás napján érvényes árfolyam)
            total_sell_huf = self._sum_in_huf(sell_rows, ccy)

            if total_sell_fc != 0:
                realized_pnl_fc = total_sell_fc - total_buy_fc
                realized_pnl_huf = total_sell_huf - total_buy_huf
                sale_date = sell_rows["Date"].max()

                realized_list.append({
                    "Ticker": ticker,
                    "Currency": ccy,
                    "Buy Sum (FC)": total_buy_fc,
                    "Sell Sum (FC)": total_sell_fc,
                    "Realized PnL (FC)": realized_pnl_fc,
                    "Buy Sum (HUF)": total_buy_huf,
                    "Sell Sum (HUF)": total_sell_huf,
                    "Realized PnL (HUF)": realized_pnl_huf,
                    "Sale Date": sale_date.strftime("%Y-%m-%d")
                })
            else:
                open_positions_list.append({
                    "Ticker": ticker,
                    "Currency": ccy,
                    "Buy Sum (FC)": total_buy_fc,
                    "Buy Sum (HUF)": total_buy_huf
                })

        realized_df = pd.DataFrame(realized_list)
        open_df = pd.DataFrame(open_positions_list)
        return realized_df, open_df

    def _sum_in_huf(self, df: pd.DataFrame, ccy: str) -> float:
        total = 0.0
        for idx, row in df.iterrows():
            rate = self.exchange_service.get_exchange_rate(row["Date"], ccy)
            if rate is None:
                rate = 0
            total += row["Net Amt."] * rate
        return total


    def _process_income(self, income_df: pd.DataFrame):
        income_rows = []
        for idx, row in income_df.iterrows():
            rate = self.exchange_service.get_exchange_rate(row["Date"], row["CCY"])
            conv_amount = row["Net Amt."] * (rate if rate is not None else 1)
            income_rows.append({
                "Date": row["Date"].strftime("%Y-%m-%d"),
                "Currency": row["CCY"],
                "Amount (FC)": row["Net Amt."],
                "Exchange Rate": rate if rate is not None else 1,
                "Amount (HUF)": conv_amount
            })
        return pd.DataFrame(income_rows)

    def to_report(self) -> dict:
        """Visszaad egy dictionary-t, melyben a Lightyear adatokhoz tartozó DataFrame-ek szerepelnek a munkalap nevekkel."""
        realized_df, open_df, interest_df, dividend_df = self.process()
        return {
            "Realizált PnL": realized_df,
            "Nyitott Pozíciók": open_df,
            "Kamat": interest_df,
            "Osztalék": dividend_df,
            "Összesítő": pd.DataFrame({
                "Category": [
                    "Realizált PnL (HUF)",
                    "Kamat (HUF)",
                    "Osztalék (HUF)",
                    "Összes bevallandó összeg (HUF)"
                ],
                "Total": [
                    realized_df["Realized PnL (HUF)"].sum() if not realized_df.empty else 0,
                    interest_df["Amount (HUF)"].sum() if not interest_df.empty else 0,
                    dividend_df["Amount (HUF)"].sum() if not dividend_df.empty else 0,
                    (realized_df["Realized PnL (HUF)"].sum() if not realized_df.empty else 0) +
                    (interest_df["Amount (HUF)"].sum() if not interest_df.empty else 0) +
                    (dividend_df["Amount (HUF)"].sum() if not dividend_df.empty else 0)
                ]
            })
        }


class ExcelReportGenerator:
    """Közös Excel jelentés generátor, amely a report_data dictionary-t várja.
       report_data: dict, ahol a kulcs a munkalap neve, az érték egy DataFrame.
    """
    def __init__(self, output_file: str = "report.xlsx"):
        self.output_file = output_file

    def apply_number_format(self, worksheet, column, start_row, format_str):
        for row in worksheet.iter_rows(min_row=start_row, min_col=column, max_col=column):
            for cell in row:
                cell.number_format = format_str

    def auto_adjust_columns(self, worksheet):
        """Automatikusan beállítja az oszlop szélességét a tartalom alapján."""
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[col_letter].width = adjusted_width

    def generate(self, report_data: dict) -> None:
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            for sheet_name, df in report_data.items():
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

            workbook = writer.book
            from excel_config import SHEET_FORMAT_CONFIGS, HUF_FORMAT, NUMBER_FORMAT
            for sheet_name, formats in SHEET_FORMAT_CONFIGS.items():
                if sheet_name in writer.sheets:
                    ws = writer.sheets[sheet_name]
                    for format_type, columns in formats.items():
                        format_str = HUF_FORMAT if format_type == "huf_format" else NUMBER_FORMAT
                        for cols in columns:
                            for col in cols:
                                self.apply_number_format(ws, col, 2, format_str)

            # Auto-adjust columns for every sheet
            for sheet in workbook.worksheets:
                self.auto_adjust_columns(sheet)

        print(f"Excel fájl sikeresen generálva: {self.output_file}")


def main():

    parser = argparse.ArgumentParser(
        description="Generál Excel jelentést a CSV fájl alapján (lightyear vagy revolut)."
    )
    # Létrehozunk egy csoportot a kötelező argumentumoknak
    required = parser.add_argument_group("required arguments")
    required.add_argument(
        "-m", "--mode",
        dest="mode",
        type=str,
        choices=["lightyear", "revolut"],
        help="A CSV fájl típusa: lightyear vagy revolut.",
        required=True,
    )
    required.add_argument(
        "-f", "--file",
        dest="filename",
        type=str,
        help="A CSV fájl elérési útja.",
        required=True,
    )

    args = parser.parse_args()
    if not os.path.exists(args.filename):
        sys.exit(f"{args.filename} file not found.")

    if args.mode.lower() == "lightyear":
        processor = LightyearProcessor(args.filename)
        report_data = processor.to_report()
        output_file = "lightyear_report.xlsx"
    # elif args.mode.lower() == "revolut":
    #     processor = RevolutProcessor(args.filename)
    #     report_data = processor.to_report()
    #     output_file = "revolut_report.xlsx"
    else:
        sys.exit("Invalid mode. Choose 'lightyear' or 'revolut'.")

    report_generator = ExcelReportGenerator(output_file)
    report_generator.generate(report_data)

if __name__ == "__main__":
    main()
