import os
import sys
import argparse
import pandas as pd
import re
import numpy as np

from datetime import datetime
from openpyxl.utils import get_column_letter
from excel_config import (
    SHEET_FORMAT_CONFIGS,
    SHEET_FORMAT_CONFIGS_REVOLUT_SAVINGS,
    HUF_FORMAT,
    NUMBER_FORMAT
)
from mnb_exchange_service import MNBExchangeService

STATEMENT_PREFIX = "[statement][transactions]"


class LightyearProcessor:
    """Lightyear CSV tranzakciós adatok feldolgozása."""
    def __init__(self, csv_file: str):
        self.csv_file = csv_file
        self.df = pd.read_csv(csv_file, parse_dates=["Date"], dayfirst=True)
        self.df["Type"] = self.df["Type"].astype(str).str.strip()
        self.df["CCY"] = self.df["CCY"].astype(str).str.strip()
        self.exchange_service = MNBExchangeService()

    def process(self):
        """Feldolgozza a CSV-t és DataFrame-eket készít:
           - realized_df: Realizált (eladott) pozíciók
           - open_df: Nyitott pozíciók
           - interest_df: Kamatjövedelem
           - dividend_df: Osztalékjövedelem
        """
        trades = self.df[
            self.df["Type"].isin(["Buy", "Sell", "Distribution"]) &
            (self.df["Ticker"].notnull()) &
            (self.df["Ticker"] != "")
        ]
        interests = self.df[self.df["Type"] == "Interest"]
        dividends = self.df[self.df["Type"] == "Dividend"]

        realized_df, open_df = self._process_trades(trades)
        interest_df = self._process_income(interests)
        dividend_df = self._process_income(dividends)
        return realized_df, open_df, interest_df, dividend_df

    def _process_trades(self, trades: pd.DataFrame):
        realized_list = []
        open_positions_list = []

        for (ticker, ccy), group in trades.groupby(["Ticker", "CCY"]):
            # Szűrés: vásárlások (Buy, Distribution) és eladások (Sell)
            buy_rows = group[group["Type"].isin(["Buy", "Distribution"])]
            sell_rows = group[group["Type"] == "Sell"]

            # Összegzés idegen pénznemben
            total_buy_fc = buy_rows["Net Amt."].sum()
            total_sell_fc = sell_rows["Net Amt."].sum() if not sell_rows.empty else 0

            # Vásárlások HUF-ban: minden sorra külön átváltás (vásárlás napján érvényes árfolyam)
            total_buy_huf = self._sum_in_huf(buy_rows, ccy)

            # Eladások HUF-ban: minden eladási sorra külön átváltás (eladás napján érvényes árfolyam)
            total_sell_huf = self._sum_in_huf(sell_rows, ccy)

            if total_sell_fc != 0:
                realized_pnl_fc = total_sell_fc - total_buy_fc
                realized_pnl_huf = total_sell_huf - total_buy_huf
                sale_date = sell_rows["Date"].max()

                realized_list.append({
                    "Ticker": ticker,
                    "Currency": ccy,
                    "Buy Sum (FC)": total_buy_fc,
                    "Sell Sum (FC)": total_sell_fc,
                    "Realized PnL (FC)": realized_pnl_fc,
                    "Buy Sum (HUF)": total_buy_huf,
                    "Sell Sum (HUF)": total_sell_huf,
                    "Realized PnL (HUF)": realized_pnl_huf,
                    "Sale Date": sale_date.strftime("%Y-%m-%d")
                })
            else:
                open_positions_list.append({
                    "Ticker": ticker,
                    "Currency": ccy,
                    "Buy Sum (FC)": total_buy_fc,
                    "Buy Sum (HUF)": total_buy_huf
                })

        realized_df = pd.DataFrame(realized_list)
        open_df = pd.DataFrame(open_positions_list)
        return realized_df, open_df

    def _sum_in_huf(self, df: pd.DataFrame, ccy: str) -> float:
        total = 0.0
        for idx, row in df.iterrows():
            rate = self.exchange_service.get_exchange_rate(row["Date"], ccy)
            if rate is None:
                rate = 0
            total += row["Net Amt."] * rate
        return total


    def _process_income(self, income_df: pd.DataFrame):
        income_rows = []
        for idx, row in income_df.iterrows():
            rate = self.exchange_service.get_exchange_rate(row["Date"], row["CCY"])
            conv_amount = row["Net Amt."] * (rate if rate is not None else 1)
            income_rows.append({
                "Date": row["Date"].strftime("%Y-%m-%d"),
                "Currency": row["CCY"],
                "Amount (FC)": row["Net Amt."],
                "Exchange Rate": rate if rate is not None else 1,
                "Amount (HUF)": conv_amount
            })
        return pd.DataFrame(income_rows)

    def to_report(self) -> dict:
        """Visszaad egy dictionary-t, melyben a Lightyear adatokhoz tartozó DataFrame-ek szerepelnek a munkalap nevekkel."""
        realized_df, open_df, interest_df, dividend_df = self.process()
        return {
            "Realizált PnL": realized_df,
            "Nyitott Pozíciók": open_df,
            "Kamat": interest_df,
            "Osztalék": dividend_df,
            "Összesítő": pd.DataFrame({
                "Category": [
                    "Realizált PnL (HUF)",
                    "Kamat (HUF)",
                    "Osztalék (HUF)",
                    "Összes bevallandó összeg (HUF)"
                ],
                "Total": [
                    realized_df["Realized PnL (HUF)"].sum() if not realized_df.empty else 0,
                    interest_df["Amount (HUF)"].sum() if not interest_df.empty else 0,
                    dividend_df["Amount (HUF)"].sum() if not dividend_df.empty else 0,
                    (realized_df["Realized PnL (HUF)"].sum() if not realized_df.empty else 0) +
                    (interest_df["Amount (HUF)"].sum() if not interest_df.empty else 0) +
                    (dividend_df["Amount (HUF)"].sum() if not dividend_df.empty else 0)
                ]
            })
        }


class RevolutProcessor:
    """Revolut CSV tranzakciós adatok feldolgozása.
       A CSV file oszlopai: Date, Ticker, Type, Quantity, Price per share, Total Amount, Currency, FX Rate.
       A tranzakciók közül a 'BUY - MARKET' és 'SELL - MARKET' típusú sorokból készíti el a realizált/nyitott pozíciókat,
       míg az 'DIVIDEND' típusú sorokból az osztalékjövedelem riportot.
    """
    def __init__(self, csv_file: str):
        self.csv_file = csv_file
        # A Date oszlop automatikus konvertálása datetime típusra (ISO 8601 formátum esetén)
        self.df = pd.read_csv(csv_file, parse_dates=["Date"])
        # Trim string mezők
        self.df["Type"] = self.df["Type"].astype(str).str.strip()
        self.df["Currency"] = self.df["Currency"].astype(str).str.strip()
        self.df["Ticker"] = self.df["Ticker"].astype(str).str.strip()
        # A "Total Amount" oszlop értékeit számmá alakítjuk (a valuta jelek eltávolításával)
        self.df["Total Amount"] = self.df["Total Amount"].apply(self._convert_currency_str)
        # Inicializáljuk az MNB árfolyam szolgáltatást
        self.exchange_service = MNBExchangeService()

    @staticmethod
    def _convert_currency_str(s):
        """Segédfüggvény: eltávolítja a valuta jeleket és egyéb nem numerikus karaktereket a sztringből,
           majd float értékké alakítja azt.
        """
        if pd.isna(s):
            return 0.0
        # Eltávolítjuk a nem számjegy, pont vagy mínusz karaktereket
        s_clean = re.sub(r"[^\d\.\-]", "", s)
        try:
            return float(s_clean)
        except Exception as e:
            print(f"Hiba a {s} érték átalakításakor: {e}")
            return 0.0

    def process(self):
        """Feldolgozza a CSV-t és DataFrame-eket készít:
           - realized_df: Realizált (eladott) pozíciók
           - open_df: Nyitott pozíciók
           - dividend_df: Osztalékjövedelem
           Az interest_df itt üres DataFrame (mivel Revolut esetén nincs kamatjövedelem).
        """
        # Tranzakciók: csak azok a sorok, ahol Ticker értékű, és a Type 'BUY - MARKET' vagy 'SELL - MARKET'
        trades = self.df[
            (self.df["Ticker"].notnull()) & (self.df["Ticker"] != "") &
            (self.df["Type"].isin(["BUY - MARKET", "SELL - MARKET"]))
        ]
        # Osztalék: a Type 'DIVIDEND'
        dividends = self.df[self.df["Type"] == "DIVIDEND"]

        realized_df, open_df = self._process_trades(trades)
        dividend_df = self._process_income(dividends)
        # Üres DataFrame az interest számára (ha nincs ilyen tétel Revolut esetén)
        interest_df = pd.DataFrame(columns=["Date", "Currency", "Amount (FC)", "Exchange Rate", "Amount (HUF)"])
        return realized_df, open_df, interest_df, dividend_df

    def _process_trades(self, trades: pd.DataFrame):
        realized_list = []
        open_positions_list = []
        # Csoportosítás Ticker és Currency szerint
        for (ticker, currency), group in trades.groupby(["Ticker", "Currency"]):
            buy_rows = group[group["Type"] == "BUY - MARKET"]
            sell_rows = group[group["Type"] == "SELL - MARKET"]

            total_buy_fc = buy_rows["Total Amount"].sum()
            total_sell_fc = sell_rows["Total Amount"].sum() if not sell_rows.empty else 0.0

            # Átváltás HUF-ra: minden sorra külön, a tranzakció napján érvényes árfolyammal
            total_buy_huf = self._sum_in_huf(buy_rows, currency)
            total_sell_huf = self._sum_in_huf(sell_rows, currency)

            if total_sell_fc != 0:
                realized_pnl_fc = total_sell_fc - total_buy_fc
                realized_pnl_huf = total_sell_huf - total_buy_huf
                sale_date = sell_rows["Date"].max()
                realized_list.append({
                    "Ticker": ticker,
                    "Currency": currency,
                    "Buy Sum (FC)": total_buy_fc,
                    "Sell Sum (FC)": total_sell_fc,
                    "Realized PnL (FC)": realized_pnl_fc,
                    "Buy Sum (HUF)": total_buy_huf,
                    "Sell Sum (HUF)": total_sell_huf,
                    "Realized PnL (HUF)": realized_pnl_huf,
                    "Sale Date": sale_date.strftime("%Y-%m-%d")
                })
            else:
                open_positions_list.append({
                    "Ticker": ticker,
                    "Currency": currency,
                    "Buy Sum (FC)": total_buy_fc,
                    "Buy Sum (HUF)": total_buy_huf
                })

        realized_df = pd.DataFrame(realized_list)
        open_df = pd.DataFrame(open_positions_list)
        return realized_df, open_df

    def _sum_in_huf(self, df: pd.DataFrame, currency: str) -> float:
        """Összegzi az adott sorokban szereplő 'Total Amount' értékeket, átváltva az aktuális árfolyammal HUF-ra."""
        total = 0.0
        for idx, row in df.iterrows():
            rate = self.exchange_service.get_exchange_rate(row["Date"], currency)
            if rate is None:
                rate = 0.0
            total += row["Total Amount"] * rate
        return total

    def _process_income(self, income_df: pd.DataFrame):
        """Feldolgozza az osztalék sorokat, átváltva az összegeket HUF-ra."""
        income_rows = []
        for idx, row in income_df.iterrows():
            rate = self.exchange_service.get_exchange_rate(row["Date"], row["Currency"])
            conv_amount = row["Total Amount"] * (rate if rate is not None else 1)
            income_rows.append({
                "Date": row["Date"].strftime("%Y-%m-%d"),
                "Currency": row["Currency"],
                "Amount (FC)": row["Total Amount"],
                "Exchange Rate": rate if rate is not None else 1,
                "Amount (HUF)": conv_amount
            })
        return pd.DataFrame(income_rows)

    def to_report(self) -> dict:
        """Visszaad egy dictionary-t, melyben a Revolut adatokhoz tartozó DataFrame-ek szerepelnek a munkalap nevekkel.
           A kulcsok:
             - 'Realizált PnL'
             - 'Nyitott Pozíciók'
             - 'Osztalék'
             - 'Összesítő'
        """
        realized_df, open_df, interest_df, dividend_df = self.process()
        summary_df = pd.DataFrame({
            "Category": [
                "Realizált PnL (HUF)",
                "Osztalék (HUF)",
                "Összes bevallandó összeg (HUF)"
            ],
            "Total": [
                realized_df["Realized PnL (HUF)"].sum() if not realized_df.empty else 0,
                dividend_df["Amount (HUF)"].sum() if not dividend_df.empty else 0,
                (realized_df["Realized PnL (HUF)"].sum() if not realized_df.empty else 0) +
                (dividend_df["Amount (HUF)"].sum() if not dividend_df.empty else 0)
            ]
        })
        return {
            "Realizált PnL": realized_df,
            "Nyitott Pozíciók": open_df,
            "Osztalék": dividend_df,
            "Összesítő": summary_df
        }


class RevolutSavingsProcessor:
    """
    Revolut deviza megtakarítási számlák tranzakciós adatok feldolgozása.
    A CSV file oszlopai: Date, Description, Value, Price per share, Quantity of shares, Currency, Value_num.
    Csak azokat a tételeket veszi figyelembe, ahol a Description "Interest..." vagy "Service Fee..." szöveggel kezdődik.
    A tranzakció napján érvényes MNB árfolyam alapján kiszámolja a HUF értéket.
    """
    def __init__(self, csv_file: str):
        self.csv_file = csv_file
        self.df = pd.read_csv(csv_file, skip_blank_lines=True)
        # Ha szükséges: töröljük az esetleges ismétlődő fejléc sorokat
        self.df = self.df[self.df["Date"] != "Date"]
        # Konvertáljuk a Date oszlopot datetime típusra; formátum például: "Dec 31, 2024, 2:21:51 AM"
        self.df["Date"] = pd.to_datetime(self.df["Date"], errors="coerce")
        self.df = self.df.dropna(subset=["Date"])
        # Tisztítjuk a Description oszlopot
        self.df["Description"] = self.df["Description"].astype(str).str.strip()
        # Feltételezzük, hogy a Value oszlopban szerepel a deviza jel; hozzuk létre a Currency és a numerikus érték oszlopát:
        self.df["Currency"] = self.df["Value"].apply(self._extract_currency)
        self.df["Value_num"] = self.df["Value"].apply(self._convert_currency_value)
        self.exchange_service = MNBExchangeService()

    def _extract_currency(self, s: str) -> str:
        """Kivonja a devizakódot a Value mezőből a szimbólum alapján."""
        if isinstance(s, str):
            if "£" in s:
                return "GBP"
            elif "$" in s:
                return "USD"
            elif "€" in s or "â¬" in s:
                return "EUR"
        return None

    def _convert_currency_value(self, s: str) -> float:
        """Eltávolítja a nem numerikus karaktereket, és float értékké alakítja a sztringet."""
        if pd.isna(s):
            return 0.0
        s_clean = re.sub(r"[^\d\.\-]", "", s)
        try:
            return float(s_clean)
        except Exception as e:
            print(f"Hiba a {s} átalakításakor: {e}")
            return 0.0

    def process(self):
        """
        Szűri a CSV adatokat, hogy csak a "Interest" vagy "Service Fee" tételek maradjanak.
        Létrehozza a YearMonth oszlopot, majd kiszámolja a napi MNB árfolyam alapján a HUF értéket.
        Két DataFrame-et készít:
        - "Megtakarítás": havi bontásban, devizanemenként és Description szerint összegzett eredeti (Value_num) és HUF értékek.
        - "Összesítő": devizanemenként az összesített értékek, valamint egy új oszlopban a bruttó (Service Fee nélküli) összeget.
        """
        df_filtered = self.df[
            self.df["Description"].str.startswith("Interest", na=False) |
            self.df["Description"].str.startswith("Service Fee", na=False)
        ].copy()

        # Számoljuk ki az egyes sorok HUF értékét
        def convert_row(row):
            rate = self.exchange_service.get_exchange_rate(row["Date"], row["Currency"])
            if rate is None:
                rate = 0
            return row["Value_num"] * rate

        df_filtered["Amount_HUF"] = df_filtered.apply(convert_row, axis=1)
        df_filtered["YearMonth"] = df_filtered["Date"].dt.strftime("%Y-%m")

        # Részletes havi bontás: csoportosítunk YearMonth, Currency és Description szerint
        monthly_df = df_filtered.groupby(
            ["YearMonth", "Currency", "Description"], as_index=False
        ).agg({
            "Value_num": "sum",
            "Amount_HUF": "sum"
        })

        # Összesítő: Először külön csoportosítjuk a "Interest" és "Service Fee" tételeket
        interest_df = df_filtered[df_filtered["Description"].str.startswith("Interest", na=False)]
        fee_df = df_filtered[df_filtered["Description"].str.startswith("Service Fee", na=False)]

        interest_summary = interest_df.groupby("Currency", as_index=False).agg({
            "Value_num": "sum",
            "Amount_HUF": "sum"
        }).rename(columns={
            "Value_num": "Interest_Eredeti",
            "Amount_HUF": "Interest_HUF"
        })

        fee_summary = fee_df.groupby("Currency", as_index=False).agg({
            "Value_num": "sum",
            "Amount_HUF": "sum"
        }).rename(columns={
            "Value_num": "Fee_Eredeti",
            "Amount_HUF": "Fee_HUF"
        })

        summary_df = pd.merge(interest_summary, fee_summary, on="Currency", how="outer").fillna(0)

        # Nettó összeg: Interest + Fee (a Fee értéke várhatóan negatív)
        summary_df["Összeg (eredeti devizában)"] = summary_df["Interest_Eredeti"] + summary_df["Fee_Eredeti"]
        summary_df["Összeg (HUF)"] = summary_df["Interest_HUF"] + summary_df["Fee_HUF"]

        # Bruttó összeg: azaz amikor a Service Fee nincs levonva = csak az Interest értékek
        summary_df["Összeg bruttó (eredeti devizában)"] = summary_df["Interest_Eredeti"]
        summary_df["Összeg bruttó (HUF)"] = summary_df["Interest_HUF"]

        # Opcionálisan rendezhetjük az oszlopokat
        summary_df = summary_df[[
            "Currency",
            "Interest_Eredeti", "Fee_Eredeti", "Összeg (eredeti devizában)",
            "Interest_HUF", "Fee_HUF", "Összeg (HUF)",
            "Összeg bruttó (eredeti devizában)", "Összeg bruttó (HUF)"
        ]]

        return {"Megtakarítás": monthly_df, "Összesítő": summary_df}


    def to_report(self) -> dict:
        """Visszaad egy dictionary-t, melyben a megtakarítási számlákhoz tartozó DataFrame-ek szerepelnek a munkalap nevekkel."""
        return self.process()


class ExcelReportGenerator:
    """Közös Excel jelentés generátor, amely a report_data dictionary-t várja.
       report_data: dict, ahol a kulcs a munkalap neve, az érték egy DataFrame.
    """
    def __init__(self, output_file: str = "report.xlsx"):
        self.output_file = output_file

    def apply_number_format(self, worksheet, column, start_row, format_str):
        for row in worksheet.iter_rows(min_row=start_row, min_col=column, max_col=column):
            for cell in row:
                cell.number_format = format_str

    def auto_adjust_columns(self, worksheet):
        """Automatikusan beállítja az oszlop szélességét a tartalom alapján."""
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[col_letter].width = adjusted_width

    def generate(self, report_data: dict, sheet_format_configs=None) -> None:
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            for sheet_name, df in report_data.items():
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

            workbook = writer.book

            for sheet_name, formats in sheet_format_configs.items():
                if sheet_name in writer.sheets:
                    ws = writer.sheets[sheet_name]
                    for format_type, columns in formats.items():
                        format_str = HUF_FORMAT if format_type == "huf_format" else NUMBER_FORMAT
                        for cols in columns:
                            for col in cols:
                                self.apply_number_format(ws, col, 2, format_str)

            # Auto-adjust columns for every sheet
            for sheet in workbook.worksheets:
                self.auto_adjust_columns(sheet)

        print(f"Excel fájl sikeresen generálva: {self.output_file}")

def main():
    parser = argparse.ArgumentParser(
        description="Generál Excel jelentést a CSV fájl alapján (lightyear, revolut, revolut_saving vagy revolut_exchange)."
    )
    required = parser.add_argument_group("required arguments")
    required.add_argument(
        "-m", "--mode",
        dest="mode",
        type=str,
        choices=["lightyear", "revolut", "revolut_exchange", "revolut_saving"],
        help="A CSV fájl típusa: lightyear, revolut, revolut_saving vagy revolut_exchange.",
        required=True,
    )
    required.add_argument(
        "-f", "--file",
        dest="filename",
        type=str,
        help="A CSV fájl elérési útja.",
        required=True,
    )

    args = parser.parse_args()
    if not os.path.exists(args.filename):
        sys.exit(f"{args.filename} file not found.")

    sheet_format = None
    if args.mode.lower() == "lightyear":
        processor = LightyearProcessor(args.filename)
        report_data = processor.to_report()
        output_file = "lightyear_report.xlsx"
        sheet_format = SHEET_FORMAT_CONFIGS
    elif args.mode.lower() == "revolut":
        processor = RevolutProcessor(args.filename)
        report_data = processor.to_report()
        output_file = "revolut_report.xlsx"
        sheet_format = SHEET_FORMAT_CONFIGS
    elif args.mode.lower() == "revolut_saving":
        processor = RevolutSavingsProcessor(args.filename)
        report_data = processor.to_report()
        output_file = "revolut_saving_report.xlsx"
        sheet_format = SHEET_FORMAT_CONFIGS_REVOLUT_SAVINGS
    else:
        sys.exit("Invalid mode. Choose 'lightyear', 'revolut', 'revolut_exchange' or 'revolut_saving'.")

    report_generator = ExcelReportGenerator(output_file)
    report_generator.generate(report_data, sheet_format_configs=sheet_format)

if __name__ == "__main__":
    main()
